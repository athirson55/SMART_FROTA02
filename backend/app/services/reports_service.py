from datetime import datetime, timedelta, timezone

from fastapi import HTTPException, status
from sqlalchemy import and_, func, select
from sqlalchemy.orm import Session, joinedload

from app.models.fleet import Driver, Vehicle, VehiclePendencia
from app.models.operations import Alert, Appointment, Maintenance
from app.services.fleet_service import serialize_vehicle
from app.services.operations_service import serialize_alert

_MONTH_LABELS_PT = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
_STATUS_LABELS = {"ATIVO": "Disponíveis", "EM_ROTA": "Em rota", "MANUTENCAO": "Manutenção", "INATIVO": "Inativos"}
_STATUS_COLORS = {"ATIVO": "#16A34A", "EM_ROTA": "#2563EB", "MANUTENCAO": "#DC2626", "INATIVO": "#7B8CA0"}
_TYPE_COLORS = {"PREVENTIVA": "#2563EB", "CORRETIVA": "#DC2626", "REVISAO": "#16A34A", "EMERGENCIAL": "#D97706"}


def dashboard_report(db: Session):
    now = datetime.now(timezone.utc)
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    next_week = today_start + timedelta(days=7)

    total_vehicles = db.scalar(select(func.count()).select_from(Vehicle)) or 0
    active_vehicles = db.scalar(select(func.count()).select_from(Vehicle).where(Vehicle.status == "ATIVO")) or 0
    in_route_vehicles = db.scalar(select(func.count()).select_from(Vehicle).where(Vehicle.status == "EM_ROTA")) or 0
    in_maintenance = db.scalar(select(func.count()).select_from(Vehicle).where(Vehicle.status == "MANUTENCAO")) or 0

    total_drivers = db.scalar(select(func.count()).select_from(Driver)) or 0
    drivers_in_route = db.scalar(select(func.count()).select_from(Driver).where(Driver.status == "EM_ROTA")) or 0
    drivers_available = db.scalar(select(func.count()).select_from(Driver).where(Driver.status == "DISPONIVEL")) or 0

    maintenances_pending = db.scalar(
        select(func.count()).select_from(Maintenance).where(Maintenance.status.in_(["PENDENTE", "AGENDADA"]))
    ) or 0
    maintenances_in_progress = db.scalar(select(func.count()).select_from(Maintenance).where(Maintenance.status == "EM_ANDAMENTO")) or 0
    total_maintenance_cost = float(db.scalar(select(func.sum(Maintenance.custo))) or 0)

    total_alerts = db.scalar(select(func.count()).select_from(Alert)) or 0
    alerts_open = db.scalar(select(func.count()).select_from(Alert).where(Alert.status != "RESOLVIDO")) or 0
    alerts_critical = db.scalar(select(func.count()).select_from(Alert).where(Alert.prioridade == "CRITICO", Alert.status != "RESOLVIDO")) or 0

    appointments_upcoming = db.scalar(
        select(func.count()).select_from(Appointment).where(
            and_(
                Appointment.data >= today_start,
                Appointment.data <= next_week,
                Appointment.status.in_(["AGENDADO", "CONFIRMADO"]),
            )
        )
    ) or 0

    appointments_overdue = db.scalar(
        select(func.count()).select_from(Appointment).where(
            and_(
                Appointment.data < today_start,
                Appointment.status.in_(["AGENDADO", "CONFIRMADO"]),
            )
        )
    ) or 0

    recent_alerts = db.scalars(
        select(Alert).options(joinedload(Alert.vehicle))
        .where(Alert.status != "RESOLVIDO")
        .order_by(Alert.created_at.desc())
        .limit(10)
    ).unique().all()

    vehicles_with_open_pendencias = db.scalars(
        select(Vehicle)
        .options(joinedload(Vehicle.pendencias))
        .where(
            Vehicle.id.in_(
                select(VehiclePendencia.vehicle_id)
                .where(VehiclePendencia.resolved_at.is_(None))
                .distinct()
            )
        )
        .limit(10)
    ).unique().all()

    veiculos_com_pendencias = []
    for v in vehicles_with_open_pendencias:
        open_p = [p for p in v.pendencias if p.resolved_at is None]
        veiculos_com_pendencias.append({
            "id": v.id,
            "modelo": v.modelo,
            "placa": v.placa,
            "status": v.status,
            "pendencias": [{"slug": p.slug, "label": p.label, "detail": p.detail, "tone": p.tone} for p in open_p],
        })

    payload = {
        "veiculos": {
            "total": total_vehicles,
            "ativos": active_vehicles,
            "emRota": in_route_vehicles,
            "emManutencao": in_maintenance,
        },
        "motoristas": {
            "total": total_drivers,
            "emRota": drivers_in_route,
            "disponiveis": drivers_available,
        },
        "manutencoes": {
            "pendentes": maintenances_pending,
            "emAndamento": maintenances_in_progress,
            "custoTotal": total_maintenance_cost,
        },
        "alertas": {
            "total": total_alerts,
            "pendentes": alerts_open,
            "criticos": alerts_critical,
        },
        "agendamentos": {
            "proximos": appointments_upcoming,
            "atrasados": appointments_overdue,
        },
        "alertasRecentes": [
            {
                "id": a.id,
                "titulo": a.titulo,
                "prioridade": a.prioridade,
                "status": a.status,
                "veiculo": {"placa": a.vehicle.placa, "modelo": a.vehicle.modelo} if a.vehicle else None,
            }
            for a in recent_alerts
        ],
        "veiculosComPendencias": veiculos_com_pendencias,
        "generatedAt": now,
    }
    return payload


def fleet_report(db: Session, search: str | None = None):
    query = select(Vehicle).options(joinedload(Vehicle.motorista)).order_by(Vehicle.created_at.desc())
    if search:
        term = f"%{search.strip()}%"
        query = query.where((Vehicle.modelo.ilike(term)) | (Vehicle.placa.ilike(term)) | (Vehicle.marca.ilike(term)))
    vehicles = db.scalars(query).unique().all()
    items = []
    for vehicle in vehicles:
        pendencias = db.scalars(select(VehiclePendencia).where(VehiclePendencia.vehicle_id == vehicle.id, VehiclePendencia.resolved_at.is_(None))).all()
        items.append({
            **serialize_vehicle(vehicle),
            "pendencias": [{"slug": p.slug, "label": p.label, "detail": p.detail, "tone": p.tone} for p in pendencias],
        })
    payload = {"generatedAt": datetime.now(timezone.utc), "items": items, "summary": {"total": len(items)}}
    return payload


def complete_report(db: Session, dias: int = 30, veiculo_id: str | None = None) -> dict:
    if veiculo_id and not db.get(Vehicle, veiculo_id):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Veículo não encontrado")
    now = datetime.now(timezone.utc)
    start = now - timedelta(days=dias)

    filters = [Maintenance.data >= start]
    if veiculo_id:
        filters.append(Maintenance.vehicle_id == veiculo_id)

    total_cost = float(db.scalar(select(func.sum(Maintenance.custo)).where(*filters)) or 0)
    maintenance_count = db.scalar(select(func.count(Maintenance.id)).where(*filters)) or 0

    # Apply vehicle filter to KPIs
    vehicle_filters = [Vehicle.status == "ATIVO"]
    alert_filters = [Alert.status != "RESOLVIDO"]
    if veiculo_id:
        vehicle_filters.append(Vehicle.id == veiculo_id)
        alert_filters.append(Alert.vehicle_id == veiculo_id)

    active_vehicles = db.scalar(select(func.count()).select_from(Vehicle).where(*vehicle_filters)) or 0
    open_alerts = db.scalar(select(func.count()).select_from(Alert).where(*alert_filters)) or 0

    # Previous period for delta
    prev_start = start - timedelta(days=dias)
    prev_filters = [Maintenance.data >= prev_start, Maintenance.data < start]
    if veiculo_id:
        prev_filters.append(Maintenance.vehicle_id == veiculo_id)
    prev_cost = float(db.scalar(select(func.sum(Maintenance.custo)).where(*prev_filters)) or 0)
    prev_count = db.scalar(select(func.count(Maintenance.id)).where(*prev_filters)) or 0

    # Monthly costs — last 12 months
    monthly_costs = []
    for i in range(11, -1, -1):
        m_start = (now.replace(day=1, hour=0, minute=0, second=0, microsecond=0) - timedelta(days=i * 28)).replace(day=1)
        m_end = (m_start + timedelta(days=32)).replace(day=1)
        monthly_filters = [Maintenance.data >= m_start, Maintenance.data < m_end]
        if veiculo_id:
            monthly_filters.append(Maintenance.vehicle_id == veiculo_id)
        cost = float(db.scalar(select(func.sum(Maintenance.custo)).where(*monthly_filters)) or 0)
        monthly_costs.append({"month": _MONTH_LABELS_PT[m_start.month - 1], "year": m_start.year, "cost": cost})

    # Maintenance by type
    type_rows = db.execute(select(Maintenance.tipo, func.count(Maintenance.id).label("cnt")).where(*filters).group_by(Maintenance.tipo)).all()
    total_types = sum(r.cnt for r in type_rows) or 1
    maintenance_by_type = [
        {"label": r.tipo.capitalize(), "value": r.cnt, "percent": round(r.cnt / total_types * 100), "color": _TYPE_COLORS.get(r.tipo, "#7B8CA0")}
        for r in type_rows
    ]

    # Vehicle stats
    veh_rows = db.execute(
        select(Maintenance.vehicle_id, func.count(Maintenance.id).label("qtd"), func.sum(Maintenance.custo).label("total_cost"), func.max(Maintenance.data).label("last_maintenance"))
        .where(*filters)
        .group_by(Maintenance.vehicle_id)
        .order_by(func.sum(Maintenance.custo).desc())
        .limit(15)
    ).all()

    vehicle_stats = []
    max_cost_v = max((float(r.total_cost or 0) for r in veh_rows), default=1)
    for r in veh_rows:
        vehicle = db.get(Vehicle, r.vehicle_id)
        if not vehicle:
            continue
        alert_count = db.scalar(select(func.count()).select_from(Alert).where(Alert.vehicle_id == r.vehicle_id, Alert.status != "RESOLVIDO")) or 0
        raw_cost = float(r.total_cost or 0)
        perf = max(0, 100 - alert_count * 15 - (10 if r.qtd > 3 else 0))
        vehicle_stats.append({
            "id": vehicle.id, "modelo": vehicle.modelo, "placa": vehicle.placa,
            "qtd": r.qtd, "totalCost": raw_cost,
            "avgCost": round(raw_cost / r.qtd) if r.qtd else 0,
            "lastMaintenance": r.last_maintenance.strftime("%d %b %Y") if r.last_maintenance else "—",
            "alerts": alert_count, "performance": perf,
            "costPct": round(raw_cost / max_cost_v * 100),
        })

    # Fleet status
    status_query = select(Vehicle.status, func.count(Vehicle.id).label("cnt")).group_by(Vehicle.status)
    if veiculo_id:
        status_query = status_query.where(Vehicle.id == veiculo_id)
    status_rows = db.execute(status_query).all()
    fleet_status = [
        {"label": _STATUS_LABELS.get(r.status, r.status), "value": r.cnt, "color": _STATUS_COLORS.get(r.status, "#7B8CA0")}
        for r in status_rows if r.cnt > 0
    ]

    return {
        "kpis": {
            "totalCost": total_cost,
            "maintenanceCount": maintenance_count,
            "activeVehicles": active_vehicles,
            "openAlerts": open_alerts,
            "deltaCost": round((total_cost - prev_cost) / prev_cost * 100) if prev_cost else 0,
            "deltaCount": maintenance_count - prev_count,
        },
        "monthlyCosts": monthly_costs,
        "maintenanceByType": maintenance_by_type,
        "vehicleStats": vehicle_stats,
        "fleetStatus": fleet_status,
    }


def export_excel(db: Session, dias: int = 30, veiculo_id: str | None = None) -> bytes:
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    data = complete_report(db, dias=dias, veiculo_id=veiculo_id)
    kpis = data["kpis"]

    wb = openpyxl.Workbook()

    # ── Helpers ─────────────────────────────────────────────────────────────
    BLUE = "2563EB"
    DARK = "1E293B"
    LIGHT_GREY = "F1F5F9"
    WHITE = "FFFFFF"
    GREEN = "16A34A"
    RED = "DC2626"
    AMBER = "D97706"

    def _header_fill(color: str) -> PatternFill:
        return PatternFill("solid", fgColor=color)

    def _bold(size: int = 11, color: str = WHITE) -> Font:
        return Font(bold=True, size=size, color=color)

    def _thin_border() -> Border:
        s = Side(style="thin", color="CBD5E1")
        return Border(left=s, right=s, top=s, bottom=s)

    def _style_header_row(ws, row_idx: int, num_cols: int, bg: str = BLUE):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = _header_fill(bg)
            cell.font = _bold(10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _thin_border()

    def _style_data_row(ws, row_idx: int, num_cols: int, alt: bool = False):
        bg = LIGHT_GREY if alt else WHITE
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = _header_fill(bg)
            cell.font = Font(size=10, color=DARK)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _thin_border()

    # ── Sheet 1: Resumo ──────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Resumo"

    ws1.merge_cells("A1:D1")
    title_cell = ws1["A1"]
    title_cell.value = f"Smart Frota — Relatório de Performance (últimos {dias} dias)"
    title_cell.font = Font(bold=True, size=14, color=WHITE)
    title_cell.fill = _header_fill(BLUE)
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws1.row_dimensions[1].height = 28

    ws1.merge_cells("A2:D2")
    ws1["A2"].value = f"Gerado em: {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M')} UTC"
    ws1["A2"].font = Font(size=9, color="64748B")
    ws1["A2"].alignment = Alignment(indent=1)
    ws1.row_dimensions[2].height = 16

    headers_kpi = ["Indicador", "Valor", "Período anterior", "Variação"]
    for col, h in enumerate(headers_kpi, 1):
        ws1.cell(row=4, column=col).value = h
    _style_header_row(ws1, 4, len(headers_kpi))

    kpi_rows = [
        ("Custo total (R$)", f"R$ {kpis['totalCost']:,.2f}", "—", f"{kpis['deltaCost']:+.0f}%"),
        ("Total de manutenções", kpis["maintenanceCount"], "—", f"{kpis['deltaCount']:+d}"),
        ("Veículos disponíveis", kpis["activeVehicles"], "—", "—"),
        ("Alertas em aberto", kpis["openAlerts"], "—", "—"),
    ]
    for r_idx, row in enumerate(kpi_rows, 5):
        for c_idx, val in enumerate(row, 1):
            ws1.cell(row=r_idx, column=c_idx).value = val
        _style_data_row(ws1, r_idx, len(headers_kpi), alt=r_idx % 2 == 0)

    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 20
    ws1.column_dimensions["C"].width = 20
    ws1.column_dimensions["D"].width = 16

    # ── Sheet 2: Veículos ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Veículos")
    headers_v = ["Veículo", "Placa", "Manutenções", "Custo total (R$)", "Custo médio (R$)", "Última manutenção", "Alertas ativos", "Performance (%)"]
    for col, h in enumerate(headers_v, 1):
        ws2.cell(row=1, column=col).value = h
    _style_header_row(ws2, 1, len(headers_v))

    for r_idx, v in enumerate(data["vehicleStats"], 2):
        values = [
            v["modelo"], v["placa"], v["qtd"],
            round(v["totalCost"], 2), round(v["avgCost"], 2),
            v["lastMaintenance"], v["alerts"], v["performance"],
        ]
        for c_idx, val in enumerate(values, 1):
            ws2.cell(row=r_idx, column=c_idx).value = val
        _style_data_row(ws2, r_idx, len(headers_v), alt=r_idx % 2 == 0)
        # colour the performance cell
        perf_cell = ws2.cell(row=r_idx, column=8)
        if v["performance"] >= 85:
            perf_cell.font = Font(bold=True, size=10, color=GREEN)
        elif v["performance"] >= 65:
            perf_cell.font = Font(bold=True, size=10, color=AMBER)
        else:
            perf_cell.font = Font(bold=True, size=10, color=RED)

    col_widths_v = [22, 12, 14, 18, 18, 20, 14, 16]
    for idx, w in enumerate(col_widths_v, 1):
        ws2.column_dimensions[get_column_letter(idx)].width = w

    # ── Sheet 3: Custos mensais ──────────────────────────────────────────────
    ws3 = wb.create_sheet("Custos mensais")
    headers_m = ["Mês", "Ano", "Custo (R$)"]
    for col, h in enumerate(headers_m, 1):
        ws3.cell(row=1, column=col).value = h
    _style_header_row(ws3, 1, len(headers_m))

    for r_idx, m in enumerate(data["monthlyCosts"], 2):
        ws3.cell(row=r_idx, column=1).value = m["month"]
        ws3.cell(row=r_idx, column=2).value = m["year"]
        ws3.cell(row=r_idx, column=3).value = round(m["cost"], 2)
        _style_data_row(ws3, r_idx, 3, alt=r_idx % 2 == 0)

    ws3.column_dimensions["A"].width = 10
    ws3.column_dimensions["B"].width = 8
    ws3.column_dimensions["C"].width = 18

    # ── Sheet 4: Status da frota ─────────────────────────────────────────────
    ws4 = wb.create_sheet("Status da frota")
    headers_s = ["Status", "Quantidade", "Participação (%)"]
    for col, h in enumerate(headers_s, 1):
        ws4.cell(row=1, column=col).value = h
    _style_header_row(ws4, 1, len(headers_s))

    total_fleet = sum(s["value"] for s in data["fleetStatus"]) or 1
    for r_idx, s in enumerate(data["fleetStatus"], 2):
        ws4.cell(row=r_idx, column=1).value = s["label"]
        ws4.cell(row=r_idx, column=2).value = s["value"]
        ws4.cell(row=r_idx, column=3).value = round(s["value"] / total_fleet * 100, 1)
        _style_data_row(ws4, r_idx, 3, alt=r_idx % 2 == 0)

    ws4.column_dimensions["A"].width = 18
    ws4.column_dimensions["B"].width = 14
    ws4.column_dimensions["C"].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def costs_report(db: Session, search: str | None = None):
    query = select(Maintenance).options(joinedload(Maintenance.vehicle).joinedload(Vehicle.motorista)).order_by(Maintenance.data.desc())
    if search:
        term = f"%{search.strip()}%"
        query = query.where((Maintenance.descricao.ilike(term)) | (Maintenance.oficina.ilike(term)) | (Maintenance.mecanico.ilike(term)))
    items = []
    total = 0.0
    for maintenance in db.scalars(query).unique().all():
        total += float(maintenance.custo or 0)
        items.append({
            "id": maintenance.id,
            "veiculo": serialize_vehicle(maintenance.vehicle) if maintenance.vehicle else None,
            "descricao": maintenance.descricao,
            "custo": float(maintenance.custo or 0),
            "data": maintenance.data,
            "status": maintenance.status,
            "prioridade": maintenance.prioridade,
        })
    payload = {"generatedAt": datetime.now(timezone.utc), "items": items, "summary": {"total": total, "count": len(items)}}
    return payload
